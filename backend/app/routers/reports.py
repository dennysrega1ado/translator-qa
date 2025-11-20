from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case, distinct
from typing import List, Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from app.database import get_db
from app import models, schemas
from app.auth import get_current_active_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/", response_model=List[schemas.ExecutionReport])
async def get_reports(
    execution_id: Optional[str] = None,
    prompt_id: Optional[int] = None,
    manual_only: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # Base query grouping by execution_id and prompt_id
    query = db.query(
        models.Translation.execution_id,
        models.Translation.prompt_id,
        models.Prompt.name.label('prompt_name'),
        func.count(models.Translation.id).label('total_translations'),
        func.count(func.distinct(models.ManualScore.translation_id)).label('translations_with_manual_scores'),

        # Automated scores averages
        func.avg(models.Translation.automated_coherence).label('avg_automated_coherence'),
        func.avg(models.Translation.automated_fidelity).label('avg_automated_fidelity'),
        func.avg(models.Translation.automated_naturalness).label('avg_automated_naturalness'),
        func.avg(models.Translation.automated_overall).label('avg_automated_overall'),

        # Manual scores averages
        func.avg(models.ManualScore.coherence).label('avg_manual_coherence'),
        func.avg(models.ManualScore.fidelity).label('avg_manual_fidelity'),
        func.avg(models.ManualScore.naturalness).label('avg_manual_naturalness'),
        func.avg(models.ManualScore.overall).label('avg_manual_overall'),
    ).join(
        models.Prompt,
        models.Translation.prompt_id == models.Prompt.id
    ).outerjoin(
        models.ManualScore,
        models.Translation.id == models.ManualScore.translation_id
    )

    # Apply filters
    if execution_id:
        query = query.filter(models.Translation.execution_id == execution_id)
    if prompt_id:
        query = query.filter(models.Translation.prompt_id == prompt_id)
    if manual_only:
        query = query.filter(models.ManualScore.id.isnot(None))

    # Group by
    query = query.group_by(
        models.Translation.execution_id,
        models.Translation.prompt_id,
        models.Prompt.name
    )

    results = query.all()

    reports = []
    for result in results:
        # Calculate combined averages (prefer manual, fallback to automated)
        total = result.total_translations
        manual_count = result.translations_with_manual_scores or 0

        def combine_scores(manual_avg, auto_avg):
            if manual_avg is not None and auto_avg is not None:
                # Weighted average based on coverage
                manual_weight = manual_count / total if total > 0 else 0
                auto_weight = (total - manual_count) / total if total > 0 else 0
                return (manual_avg * manual_weight) + (auto_avg * auto_weight)
            elif manual_avg is not None:
                return manual_avg
            elif auto_avg is not None:
                return auto_avg
            return None

        report = schemas.ExecutionReport(
            execution_id=result.execution_id,
            prompt_id=result.prompt_id,
            prompt_name=result.prompt_name,
            total_translations=total,
            translations_with_manual_scores=manual_count,
            manual_score_percentage=round((manual_count / total * 100) if total > 0 else 0, 2),

            avg_automated_coherence=round(result.avg_automated_coherence, 2) if result.avg_automated_coherence else None,
            avg_automated_fidelity=round(result.avg_automated_fidelity, 2) if result.avg_automated_fidelity else None,
            avg_automated_naturalness=round(result.avg_automated_naturalness, 2) if result.avg_automated_naturalness else None,
            avg_automated_overall=round(result.avg_automated_overall, 2) if result.avg_automated_overall else None,

            avg_manual_coherence=round(result.avg_manual_coherence, 2) if result.avg_manual_coherence else None,
            avg_manual_fidelity=round(result.avg_manual_fidelity, 2) if result.avg_manual_fidelity else None,
            avg_manual_naturalness=round(result.avg_manual_naturalness, 2) if result.avg_manual_naturalness else None,
            avg_manual_overall=round(result.avg_manual_overall, 2) if result.avg_manual_overall else None,

            avg_combined_coherence=round(combine_scores(result.avg_manual_coherence, result.avg_automated_coherence), 2) if combine_scores(result.avg_manual_coherence, result.avg_automated_coherence) else None,
            avg_combined_fidelity=round(combine_scores(result.avg_manual_fidelity, result.avg_automated_fidelity), 2) if combine_scores(result.avg_manual_fidelity, result.avg_automated_fidelity) else None,
            avg_combined_naturalness=round(combine_scores(result.avg_manual_naturalness, result.avg_automated_naturalness), 2) if combine_scores(result.avg_manual_naturalness, result.avg_automated_naturalness) else None,
            avg_combined_overall=round(combine_scores(result.avg_manual_overall, result.avg_automated_overall), 2) if combine_scores(result.avg_manual_overall, result.avg_automated_overall) else None,
        )
        reports.append(report)

    return reports


@router.get("/summary", response_model=schemas.SummaryReport)
async def get_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    """
    Get overall quality summary based on manual scores from all users
    """
    # Total translations
    total_translations = db.query(func.count(models.Translation.id)).scalar()

    # Translations with at least one manual score
    translations_reviewed = db.query(
        func.count(distinct(models.ManualScore.translation_id))
    ).scalar() or 0

    # Calculate review percentage
    review_percentage = round((translations_reviewed / total_translations * 100) if total_translations > 0 else 0, 2)

    # Average manual scores across all translations
    avg_scores = db.query(
        func.avg(models.ManualScore.overall).label('avg_overall'),
        func.avg(models.ManualScore.coherence).label('avg_coherence'),
        func.avg(models.ManualScore.fidelity).label('avg_fidelity'),
        func.avg(models.ManualScore.naturalness).label('avg_naturalness'),
    ).first()

    # Get contributors (users who have submitted manual scores)
    contributors_query = db.query(
        models.User.username,
        func.count(models.ManualScore.id).label('contributions')
    ).join(
        models.ManualScore,
        models.User.id == models.ManualScore.user_id
    ).group_by(
        models.User.id,
        models.User.username
    ).order_by(
        func.count(models.ManualScore.id).desc()
    ).all()

    contributors = [
        schemas.ContributorUser(username=c.username, contributions=c.contributions)
        for c in contributors_query
    ]

    return schemas.SummaryReport(
        total_translations=total_translations,
        translations_reviewed=translations_reviewed,
        review_percentage=review_percentage,
        avg_manual_overall=round(avg_scores.avg_overall, 3) if avg_scores.avg_overall else None,
        avg_manual_coherence=round(avg_scores.avg_coherence, 3) if avg_scores.avg_coherence else None,
        avg_manual_fidelity=round(avg_scores.avg_fidelity, 3) if avg_scores.avg_fidelity else None,
        avg_manual_naturalness=round(avg_scores.avg_naturalness, 3) if avg_scores.avg_naturalness else None,
        contributors=contributors
    )


@router.get("/export/user-reviews")
async def export_user_reviews(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    """
    Export current user's translation reviews to Excel
    Excludes execution_id and prompt_id fields
    """
    # Query translations with manual scores from current user
    translations_query = db.query(
        models.Translation,
        models.ManualScore,
        models.Prompt.name.label('prompt_name')
    ).join(
        models.ManualScore,
        models.Translation.id == models.ManualScore.translation_id
    ).join(
        models.Prompt,
        models.Translation.prompt_id == models.Prompt.id
    ).filter(
        models.ManualScore.user_id == current_user.id
    ).order_by(
        models.Translation.created_at.desc()
    ).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "My Reviews"

    # Define headers (excluding execution_id and prompt_id)
    headers = [
        'ID',
        'Prompt Name',
        'Source Language',
        'Target Language',
        'Original Content',
        'Translated Content',
        'Automated Coherence',
        'Automated Fidelity',
        'Automated Naturalness',
        'Automated Overall',
        'My Coherence',
        'My Fidelity',
        'My Naturalness',
        'My Overall',
        'My Notes',
        'Review Date',
        'Translation Date'
    ]

    # Style header row
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add data rows
    for row_num, (translation, manual_score, prompt_name) in enumerate(translations_query, 2):
        ws.cell(row=row_num, column=1, value=translation.id)
        ws.cell(row=row_num, column=2, value=prompt_name)
        ws.cell(row=row_num, column=3, value=translation.source_language)
        ws.cell(row=row_num, column=4, value=translation.target_language)
        ws.cell(row=row_num, column=5, value=translation.original_content)
        ws.cell(row=row_num, column=6, value=translation.translated_content)

        # Automated scores (convert to 0-10 scale for readability)
        ws.cell(row=row_num, column=7, value=round(translation.automated_coherence * 10, 2) if translation.automated_coherence else None)
        ws.cell(row=row_num, column=8, value=round(translation.automated_fidelity * 10, 2) if translation.automated_fidelity else None)
        ws.cell(row=row_num, column=9, value=round(translation.automated_naturalness * 10, 2) if translation.automated_naturalness else None)
        ws.cell(row=row_num, column=10, value=round(translation.automated_overall * 10, 2) if translation.automated_overall else None)

        # Manual scores (convert to 0-10 scale for readability)
        ws.cell(row=row_num, column=11, value=round(manual_score.coherence * 10, 2) if manual_score.coherence else None)
        ws.cell(row=row_num, column=12, value=round(manual_score.fidelity * 10, 2) if manual_score.fidelity else None)
        ws.cell(row=row_num, column=13, value=round(manual_score.naturalness * 10, 2) if manual_score.naturalness else None)
        ws.cell(row=row_num, column=14, value=round(manual_score.overall * 10, 2) if manual_score.overall else None)

        ws.cell(row=row_num, column=15, value=manual_score.notes)
        ws.cell(row=row_num, column=16, value=manual_score.created_at.strftime('%Y-%m-%d %H:%M:%S') if manual_score.created_at else None)
        ws.cell(row=row_num, column=17, value=translation.created_at.strftime('%Y-%m-%d %H:%M:%S') if translation.created_at else None)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename with timestamp
    filename = f"my_reviews_{current_user.username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
